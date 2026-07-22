#!/usr/bin/env python3
"""周度数据复盘 —— 读取最新Excel，生成复盘报告到 08_数据复盘/"""

import warnings
warnings.filterwarnings("ignore")

import openpyxl, glob, os, json
from datetime import datetime

VAULT = os.path.expanduser("~/Documents/Obsidian/PersonalGrowth_AI_Vault")
OUT_DIR = os.path.join(VAULT, "08_数据复盘")

files = sorted(glob.glob(os.path.expanduser("~/Downloads/笔记列表明细表*.xlsx")),
               key=os.path.getmtime, reverse=True)

if not files:
    print("❌ 未找到数据文件")
    exit(1)

wb = openpyxl.load_workbook(files[0], data_only=True)
ws = wb[wb.sheetnames[0]]
rows = list(ws.iter_rows(values_only=True))

today = datetime.now().strftime("%Y-%m-%d")
source_file = os.path.basename(files[0])

# Parse all content
contents = []
for r in rows[2:]:
    if r[0] and int(r[3] or 0) > 0:
        exp = int(r[3] or 0)
        contents.append({
            "title": str(r[0]).strip(),
            "date_str": str(r[1])[:10] if r[1] else "",
            "exp": exp,
            "views": int(r[4] or 0),
            "ctr": float(r[5] or 0),
            "likes": int(r[6] or 0),
            "comments": int(r[7] or 0),
            "saves": int(r[8] or 0),
            "fans": int(r[9] or 0),
            "shares": int(r[10] or 0),
            "avg_read": int(r[11] or 0),
        })

# Stats
total_exp = sum(c["exp"] for c in contents)
total_views = sum(c["views"] for c in contents)
total_interaction = sum(c["likes"] + c["comments"] + c["saves"] + c["shares"] for c in contents)
total_fans = sum(c["fans"] for c in contents)
count = len(contents)

# Sort by exposure
contents.sort(key=lambda x: x["exp"], reverse=True)

# Identify INFP vs others
infp_keywords = ["INFP", "内耗", "高敏感", "莫向外求"]
infp_contents = [c for c in contents if any(k in c["title"] for k in infp_keywords)]
other_contents = [c for c in contents if c not in infp_contents]

infp_exp = sum(c["exp"] for c in infp_contents)
infp_interaction = sum(c["likes"] + c["comments"] + c["saves"] + c["shares"] for c in infp_contents)

# Top 3
top3 = contents[:3]

# Generate report
report = f"""---
title: {today} 周度数据复盘
date: {today}
agent: 复盘Agent (自动化)
status: 已完成
tags: [数据复盘, 自动化报告]
---

# {today} 周度数据复盘

> 🤖 自动化生成 · 数据源: {source_file}
> 总计 {count} 篇 | 总曝光: {total_exp:,} | 总互动: {total_interaction}

---

## 📊 核心指标

| 指标 | 数值 |
|------|:----:|
| 内容总数 | {count} 篇 |
| 总曝光 | **{total_exp:,}** |
| 总观看 | **{total_views:,}** |
| 总互动 | **{total_interaction}** |
| 平均互动率 | **{round(total_interaction/max(total_views,1)*100,1)}%** |
| 总涨粉 | {total_fans} |

## 🔥 Top 3 内容

| # | 标题 | 曝光 | CTR | 互动 |
|:-:|------|:----:|:---:|:----:|
"""

for i, c in enumerate(top3):
    flag = "🏆" if i == 0 else ("🥈" if i == 1 else "🥉")
    interaction = c["likes"] + c["comments"] + c["saves"] + c["shares"]
    title_short = c["title"][:30]
    report += f"| {flag} | {title_short} | {c['exp']:,} | {c['ctr']*100:.1f}% | {interaction} |\n"

report += f"""
## 🎯 INFP系列表现

| 指标 | 数值 |
|------|:----:|
| INFP系列篇数 | {len(infp_contents)} |
| INFP系列曝光 | **{infp_exp:,}** |
| INFP系列互动 | **{infp_interaction}** |
| 占总互动比 | **{round(infp_interaction/max(total_interaction,1)*100,1)}%** |

## 📋 全部内容明细

| 标题 | 曝光 | CTR | 互动 | 互动率 |
|------|:----:|:---:|:----:|:------:|
"""

for c in contents:
    interaction = c["likes"] + c["comments"] + c["saves"] + c["shares"]
    rate = round(interaction / max(c["views"], 1) * 100, 1)
    flag = "🔥" if c["exp"] > 10000 else ("✅" if c["exp"] > 1000 else "⚡")
    title_short = c["title"][:25]
    report += f"| {flag} {title_short} | {c['exp']:,} | {c['ctr']*100:.1f}% | {interaction} | {rate}% |\n"

report += f"""
---

> 🤖 自动化生成 · 点击按钮 [🔄 刷新] 可重新生成
> 深度分析请切换到 Claude Code: "对最新数据做深度复盘"
"""

# Save
os.makedirs(OUT_DIR, exist_ok=True)
filename = f"{today}_周度数据复盘_auto.md"
filepath = os.path.join(OUT_DIR, filename)
with open(filepath, "w", encoding="utf-8") as f:
    f.write(report)

print(f"✅ 复盘报告已生成")
print(f"📄 {filepath}")
print(f"📊 {count}篇 | 曝光:{total_exp:,} | 互动:{total_interaction}")
# Output the vault-relative path for the open command
print(f"FILE:{filepath}")
